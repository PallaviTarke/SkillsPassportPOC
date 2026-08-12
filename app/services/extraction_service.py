"""
engine1.py — Engine 1: Curriculum Skill Extractor  (UPDATED)

Key changes vs original:
  • EXTRACTION_PROMPT: new STEP 4 forces granular, tool-specific skill extraction
    (e.g. "AWS Lambda", "OAuth 2.0", "Vertex AI") instead of vague rolled-up phrases.
    Adds `raw_skill_keywords` field for tool names mentioned in the unit.
  • LIGHTCAST_BATCH_PROMPT: returns UP TO 3 LC skills per raw skill (array), with
    hard negative rules:  no certifications, no vendor-assumed mappings.
  • map_skills_to_lightcast / merge logic updated to handle the new array shape.
  • Excel builder updated: lightcast_skills column shows all matched names joined.
"""

from __future__ import annotations

import io
import json
import re
import traceback
from collections import Counter
from concurrent.futures import as_completed
from datetime import datetime, timezone
from threading import Lock

import fitz
import pdfplumber
from google.cloud import documentai_v1 as documentai
from google.oauth2 import service_account
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import openpyxl

from google.genai import types as genai_types

from app.core.config import (
    GEMINI_MODEL, GEMINI_API_KEY,
    DOCAI_PROJECT_ID, DOCAI_LOCATION, DOCAI_PROCESSOR_ID,
    KEY_PATH, OPENAI_API_KEY,
    get_genai_client, _gemini_with_retry,
    _executor, _job_store, _result_store, _result_lock,
    sse_msg, log, _call_gemini_json,
)
from app.core.database import (
    save_extracted_to_mongo,
    generate_openai_embeddings,
    _get_lightcast_docs,
    _cosine_top_k,
)

# ══════════════════════════════════════════════════════════════════════════════
# PDF EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

MAX_DOCAI_PAGES = 15

def split_pdf_bytes(pdf_bytes, chunk_size=MAX_DOCAI_PAGES):
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    total  = len(doc)
    result = []
    for start in range(0, total, chunk_size):
        end = min(start + chunk_size, total)
        sub = fitz.open()
        sub.insert_pdf(doc, from_page=start, to_page=end - 1)
        buf = io.BytesIO()
        sub.save(buf)
        sub.close()
        result.append((buf.getvalue(), start, end - 1))
    doc.close()
    return result, total

def get_page_count(pdf_bytes):
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    n = len(doc)
    doc.close()
    return n

_docai_client = None

def _build_docai_client():
    global _docai_client
    if _docai_client is not None:
        return _docai_client
    if KEY_PATH.exists():
        creds = service_account.Credentials.from_service_account_file(
            str(KEY_PATH), scopes=["https://www.googleapis.com/auth/cloud-platform"]
        )
        _docai_client = documentai.DocumentProcessorServiceClient(credentials=creds)
    else:
        _docai_client = documentai.DocumentProcessorServiceClient()
    return _docai_client

def _extract_chunk_docai(chunk_bytes: bytes) -> str:
    client = _build_docai_client()
    name   = client.processor_path(DOCAI_PROJECT_ID, DOCAI_LOCATION, DOCAI_PROCESSOR_ID)
    result = client.process_document(request=documentai.ProcessRequest(
        name=name,
        raw_document=documentai.RawDocument(content=chunk_bytes, mime_type="application/pdf"),
    ))
    doc       = result.document
    full_text = doc.text or ""

    table_chunks = []
    for p_num, page in enumerate(doc.pages, 1):
        for t_idx, table in enumerate(page.tables):
            lines = [f"\n[Table {t_idx+1} — Page {p_num}]"]
            for hr in table.header_rows:
                lines.append("HEADER: " + " | ".join(_cell(c, full_text) for c in hr.cells))
            for row in table.body_rows:
                lines.append(" | ".join(_cell(c, full_text) for c in row.cells))
            table_chunks.append("\n".join(lines))

    if table_chunks:
        full_text += "\n\n=== EXTRACTED TABLES ===\n" + "\n".join(table_chunks)
    return full_text

def _cell(cell, full_text: str) -> str:
    parts = []
    for seg in cell.layout.text_anchor.text_segments:
        s = int(seg.start_index) if seg.start_index else 0
        e = int(seg.end_index)   if seg.end_index   else 0
        parts.append(full_text[s:e].strip())
    return " ".join(parts).strip()

def extract_text_pymupdf(pdf_bytes):
    doc   = fitz.open(stream=pdf_bytes, filetype="pdf")
    pages = []
    for page in doc:
        blocks = page.get_text("blocks")
        blocks.sort(key=lambda b: (round(b[1] / 20), b[0]))
        pages.append("\n".join(b[4].strip() for b in blocks if b[4].strip()))
    doc.close()
    return "\n\n".join(pages)

def extract_text_pdfplumber(pdf_bytes):
    parts = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            for table in (page.extract_tables() or []):
                for row in table:
                    if row:
                        parts.append(" | ".join(str(c).strip() for c in row if c))
            t = page.extract_text()
            if t:
                parts.append(t)
    return "\n".join(parts)

def extract_full_text(pdf_bytes: bytes, emit=None) -> tuple[str, str]:
    """Extract text from PDF using Document AI (chunked) → PyMuPDF → pdfplumber."""
    def _emit(event, data):
        if emit:
            emit(event, data)

    total_pages = get_page_count(pdf_bytes)

    if all([DOCAI_PROJECT_ID, DOCAI_PROCESSOR_ID]):
        try:
            _emit("progress", {"step": "extract",
                "msg": f"Document AI: processing {total_pages} pages in chunks…"})
            chunks, _ = split_pdf_bytes(pdf_bytes)
            chunk_texts = [None] * len(chunks)
            lock = Lock()
            done = [0]

            def proc(idx, cb, p0, p1):
                text = _extract_chunk_docai(cb)
                with lock:
                    done[0] += 1
                    _emit("progress", {"step": "extract",
                        "msg": f"Document AI: pages {p0+1}–{p1+1} ({done[0]}/{len(chunks)} chunks)",
                        "pct": int(done[0] / len(chunks) * 40)})
                return idx, text

            futs = [_executor.submit(proc, i, cb, p0, p1) for i, (cb, p0, p1) in enumerate(chunks)]
            for fut in as_completed(futs):
                idx, text = fut.result()
                chunk_texts[idx] = text

            full = "\n".join(t for t in chunk_texts if t)
            return full, f"Document AI ({total_pages} pages, {len(chunks)} chunks)"
        except Exception as e:
            _emit("progress", {"step": "extract", "msg": f"Document AI failed ({e}), falling back…"})

    try:
        _emit("progress", {"step": "extract",
            "msg": f"PyMuPDF: extracting {total_pages} pages…"})
        text = extract_text_pymupdf(pdf_bytes)
        if text and len(text.strip()) > 200:
            _emit("progress", {"step": "extract",
                "msg": f"PyMuPDF: {len(text):,} chars.", "pct": 40})
            return text, f"PyMuPDF ({total_pages} pages)"
    except Exception as e:
        _emit("progress", {"step": "extract", "msg": f"PyMuPDF failed ({e}), trying pdfplumber…"})

    _emit("progress", {"step": "extract", "msg": "pdfplumber: extracting with table detection…"})
    text = extract_text_pdfplumber(pdf_bytes)
    _emit("progress", {"step": "extract", "msg": f"pdfplumber: {len(text):,} chars.", "pct": 40})
    return text, f"pdfplumber ({total_pages} pages)"

CHUNK_SIZE    = 60_000
CHUNK_OVERLAP = 3_000

def split_text_into_chunks(text):
    if len(text) <= CHUNK_SIZE:
        return [text]
    chunks = []
    pos = 0
    while pos < len(text):
        end = pos + CHUNK_SIZE
        if end >= len(text):
            chunks.append(text[pos:])
            break
        boundary = text.rfind("\n\n", pos + CHUNK_SIZE // 2, end)
        if boundary == -1:
            boundary = text.rfind("\n", pos + CHUNK_SIZE // 2, end)
        if boundary == -1:
            boundary = end
        chunks.append(text[pos: boundary + 1])
        pos = max(boundary - CHUNK_OVERLAP, pos + 1)
    return chunks


# ══════════════════════════════════════════════════════════════════════════════
# SYSTEM PROMPT (unchanged)
# ══════════════════════════════════════════════════════════════════════════════

EXTRACTION_SYSTEM_PROMPT = """You are a senior curriculum analyst and industry skills taxonomist with dual expertise:
  (1) Academic curriculum design — Bloom's taxonomy, credit frameworks, Indian university syllabus formats
      (MU, SPPU, VTU, GTU, JNTU, Anna University, CBCS, NEP 2020, OBE patterns)
  (2) Industry talent intelligence — skill naming on job descriptions, LinkedIn, ATS systems

Your mission: transform raw syllabus content into a HYBRID SKILL MAP that is:
  • Academically faithful  — grounded in what the syllabus actually teaches
  • Industry-searchable    — using terms hiring managers and ATS systems recognise
  • Evidence-based         — every proficiency decision cites the signals you actually found

FUNDAMENTAL RULE — WORK WITH WHAT EXISTS:
  Syllabi vary enormously. Some have full OBE blocks (COs, Bloom\'s levels, L:T:P, exam marks).
  Others have only a subject name and a list of topics. Many fall somewhere in between.
  You MUST produce a useful output regardless of how sparse or rich the metadata is.
  Never fail or skip a skill because one specific piece of information is absent.
  Always use whatever IS present and honestly report what you relied on.
"""

# ══════════════════════════════════════════════════════════════════════════════
# ★ UPDATED EXTRACTION PROMPT
# ══════════════════════════════════════════════════════════════════════════════

EXTRACTION_PROMPT = """Extract a comprehensive, evidence-based, hybrid SKILL MAP from the curriculum below.

  Department : {department}
  Year       : {year}
  Semester   : {semester}

════════════════════════════════════════════════════════════════
STEP 1 — COLLECT AVAILABLE SIGNALS (use only what exists)
════════════════════════════════════════════════════════════════

For each subject, scan the text and collect whichever of these signals are present.
SKIP any signal that is simply not there — do not invent or assume it.

  SIGNAL A — Course Outcomes (COs) with explicit Bloom\'s Cognitive Domain label
    e.g.  "CO4  Analyze  Compare various channels of social media…"
    If present: extract the Bloom\'s verb (Remember / Understand / Apply /
    Analyze / Evaluate / Create) assigned to each CO.

  SIGNAL B — Course Outcomes WITHOUT a Bloom\'s label
    If present: infer the Bloom\'s level from the CO\'s action verb:
      define / list / recall / name           → Remember   (L1)
      explain / describe / summarise / classify → Understand (L2)
      use / demonstrate / implement / solve   → Apply      (L3)
      compare / differentiate / examine       → Analyze    (L4)
      judge / justify / critique / evaluate   → Evaluate   (L5)
      design / construct / develop / create   → Create     (L6)

  SIGNAL C — Course Objectives / Aim (corroborating signal only)
  SIGNAL D — Credits / L:T:P  [METADATA ONLY — do NOT use for proficiency]
  SIGNAL E — Contact Hours (≤30 hrs → Beginner · 31–60 → Intermediate · >60 → Advanced)
  SIGNAL F — Examination Scheme (External ratio > 60% → lean one level higher)
  SIGNAL G — Prerequisites (listed → students already have prior knowledge → shift up)
  SIGNAL H — Unit / Module topics (vocabulary signals depth)
  SIGNAL I — Subject name / code only (last resort)

════════════════════════════════════════════════════════════════
STEP 2 — PROFICIENCY DECISION: GRACEFUL FALLBACK CHAIN
════════════════════════════════════════════════════════════════

Work down this chain and STOP at the first level that yields useful evidence.
⚠ NEVER use credit count (Signal D) in this decision.

  Level 1 — Explicit Bloom\'s label on COs (Signal A)
    Bloom\'s → Proficiency:
      Remember (L1) · Understand (L2)  →  Beginner
      Apply    (L3) · Analyze    (L4)  →  Intermediate
      Evaluate (L5) · Create     (L6)  →  Advanced

  Level 2 — COs without Bloom\'s label (Signal B — infer from verb)
  Level 3 — Objective verbs + Contact Hours (Signals C, E)
  Level 4 — Exam scheme + Prerequisites (Signals F, G)
  Level 5 — Topic vocabulary (Signal H)
  Level 6 — Subject name / semester position (Signal I, lowest confidence)

PROFICIENCY DEFINITIONS:
  Beginner     = recall, awareness, definition, conceptual overview, first exposure
  Intermediate = application, implementation, analysis, comparison, problem-solving
  Advanced     = design, evaluation, synthesis, research-level, cross-domain integration

════════════════════════════════════════════════════════════════
STEP 3 — SKILL NAMING: GRANULAR OVER GENERIC
════════════════════════════════════════════════════════════════

⚠ PRIMARY RULE: Extract SPECIFIC, GRANULAR skills — not vague umbrella terms.
   A student\'s skill passport must list WHAT they actually learned,
   not a catch-all phrase that could mean anything.

GRANULARITY DECISION TABLE:
┌──────────────────────────────────────────────────────────────────────────────┐
│ Syllabus says…                         │ Extract as…                        │
├──────────────────────────────────────────────────────────────────────────────┤
│ "OAuth 2.0, JWT, API Keys"             │ "API Authentication" (ONE skill;   │
│                                        │  raw_skill_keywords: ["OAuth 2.0", │
│                                        │  "JWT", "API Keys"])               │
├──────────────────────────────────────────────────────────────────────────────┤
│ "AWS Lambda, Azure Functions,          │ "Serverless Functions" (ONE skill; │
│  Google Cloud Functions"               │  raw_skill_keywords: ["AWS Lambda",│
│                                        │  "Azure Functions","GCF"])         │
├──────────────────────────────────────────────────────────────────────────────┤
│ "AWS Rekognition, Google Vision AI,    │ "Computer Vision APIs" (ONE skill; │
│  Azure Computer Vision"                │  raw_skill_keywords: ["AWS         │
│                                        │  Rekognition","Google Vision AI",  │
│                                        │  "Azure Computer Vision"])         │
├──────────────────────────────────────────────────────────────────────────────┤
│ "Kruskal\'s algorithm, Prim\'s algorithm" │ TWO skills: "Minimum Spanning   │
│  (different algorithmic concepts)      │  Tree Algorithms" AND              │
│                                        │  raw_skill_keywords per skill      │
├──────────────────────────────────────────────────────────────────────────────┤
│ "Introduction to Cloud APIs,           │ THREE skills: "RESTful API Design",│
│  RESTful, SOAP, GraphQL, WebSockets"   │ "GraphQL", "WebSocket APIs"        │
│                                        │  NOT one "Cloud API Concepts"      │
├──────────────────────────────────────────────────────────────────────────────┤
│ "Heap Sort, Quicksort"                 │ ONE skill: "Sorting Algorithms"    │
│  (variants of the same concept)        │  raw_skill_keywords: ["Heap Sort", │
│                                        │  "Quicksort"]                      │
└──────────────────────────────────────────────────────────────────────────────┘

CONSOLIDATION RULES (when to merge vs split):
  MERGE if: topics are sub-techniques of the same skill (sorting variants, auth mechanisms)
  SPLIT if: topics are conceptually distinct skills (graph algorithms vs DP vs backtracking)
  ALWAYS SPLIT platform-agnostic skills from platform-specific lab work if both exist.

NAMING RULES:
  1. Noun phrase, 2–5 words. NO verbs.
  2. Use the most specific, industry-recognisable term.
  3. If a specific technology name IS the industry skill (e.g. "Dynamic Programming",
     "OAuth 2.0", "Dijkstra\'s Algorithm"), use it directly as the skill name.
  4. Do NOT invent platform assumptions:
     — If the syllabus says "Cloud Functions" without specifying AWS/Azure/GCP,
       do NOT write "AWS Lambda". Write "Serverless Computing" or "Cloud Functions".
     — If the syllabus covers all three cloud providers equally, do NOT pick one.

WRONG EXAMPLES:
  ✗ "Cloud API Concepts"   (too vague — what API? what concept?)
  ✗ "Algorithm Analysis"   (too vague — which algorithms?)
  ✗ "Cloud Services"       (meaningless umbrella)
  ✗ "Problem Solving"      (generic non-domain)

RIGHT EXAMPLES:
  ✓ "RESTful API Design"
  ✓ "Dynamic Programming"
  ✓ "OAuth 2.0 Authentication"
  ✓ "Greedy Algorithms"
  ✓ "Serverless Computing"    (if all three providers are covered equally)
  ✓ "Computer Vision APIs"    (if AWS/GCP/Azure vision APIs are all covered)
  ✓ "Search Engine Optimization"
  ✓ "Social Media Analytics"
  ✓ "Graph Traversal Algorithms"
  ✓ "NP-Completeness Theory"

════════════════════════════════════════════════════════════════
STEP 4 — UNIT-LEVEL SCAN: EXTRACT ALL NAMED TECHNOLOGIES
════════════════════════════════════════════════════════════════

For EVERY unit / module in each subject:
  (a) List all NAMED technologies, tools, APIs, frameworks, algorithms, or standards.
      E.g.: "AWS Lambda", "Dijkstra\'s Algorithm", "OAuth 2.0", "Google Pub/Sub",
            "Floyd-Warshall", "GDPR", "Gemini", "LangChain", "Neo4j", etc.
  (b) Decide if each named item is:
      — A SKILL ITSELF (e.g. "Dynamic Programming", "Docker", "OAuth 2.0")
        → create a separate skill row for it, Bloom\'s from the CO that maps this unit
      — A TOOL implementing a broader skill (e.g. "AWS Rekognition" implements "Computer Vision")
        → add it to raw_skill_keywords of the broader skill row
  (c) NEVER silently drop a named technology. It either becomes a skill or a keyword.

════════════════════════════════════════════════════════════════
STEP 5 — PRACTICAL / LAB SKILLS
════════════════════════════════════════════════════════════════

If ANY of these indicate a practical component, extract lab skills separately:
  • L:T:P shows P > 0  (e.g. L:T:P = 3:0:2)
  • Text mentions "lab", "practical", "hands-on", "workshop", "exercises"
  • Exam scheme includes "Practical" or "Oral" marks
Lab skills are typically Apply level (L3) → Intermediate unless evidence says otherwise.

════════════════════════════════════════════════════════════════
OUTPUT FORMAT
════════════════════════════════════════════════════════════════

Return ONLY a valid JSON array. No markdown fences. No preamble. No explanation.

FIELD RULES:
  skill               → specific noun phrase (see Step 3). NOT a vague umbrella term.
  raw_skill_keywords  → list of specific tool/API/algorithm names found in the unit
                        that underpin this skill. Empty list [] if none named.
                        e.g. ["AWS Lambda", "Google Cloud Functions", "Azure Functions"]
  bloom_level         → explicit or inferred Bloom\'s level if COs exist; OR "" if no COs
  bloom_explanation   → 7-10 word explanation of what this Bloom\'s level means for this skill.
                        Explain the cognitive level and how it applies to the skill.
                        e.g. "Apply level: student implements cloud deployment solutions"
                        OR "" if no Bloom\'s level assigned
  proficiency_rationale → name the EXACT signals used, e.g.:
                          "COs explicit (Analyze L4) → Intermediate; corroborated 45 hrs"

[
  {{
    "department":            "{department}",
    "year":                  "{year}",
    "semester":              "{semester}",
    "subject":               "<exact subject code + name as in syllabus>",
    "credits":               "<credit string or empty>",
    "total_hours":           "<hours string or empty>",
    "bloom_level":           "<Remember|Understand|Apply|Analyze|Evaluate|Create or empty>",
    "bloom_explanation":     "<7-10 words explaining this level for this skill, or empty>",
    "skill":                 "<specific skill noun phrase — see Step 3>",
    "raw_skill_keywords":    ["<specific tool/API/algorithm name>", "..."],
    "proficiency":           "Beginner|Intermediate|Advanced",
    "proficiency_rationale": "<which signals were used and what they said>"
  }}
]

════════════════════════════════════════════════════════════════
CURRICULUM TEXT (chunk {chunk_idx} of {total_chunks})
════════════════════════════════════════════════════════════════
{text}
"""


# ══════════════════════════════════════════════════════════════════════════════
# ★ UPDATED LIGHTCAST BATCH PROMPT — multiple matches, hard negative rules
# ══════════════════════════════════════════════════════════════════════════════

LIGHTCAST_BATCH_PROMPT = """You are a skill taxonomy expert. For each curriculum skill below,
pick EXACTLY ONE best-matching LightCast candidate.

════════════════════
HARD NEGATIVE RULES
════════════════════
1. NEVER pick a CERTIFICATION skill unless the syllabus explicitly says students earn it.
   ✗ "SEO Certification", "AWS Certified" → pick "Search Engine Optimization", "Amazon Web Services"

2. NEVER pick a VENDOR-SPECIFIC skill unless that vendor is explicitly named in the syllabus.
   ✗ Syllabus says "Cloud Functions" (no vendor) → pick "Serverless Computing", not "AWS Lambda"
   ✓ Only pick "AWS Lambda" if the syllabus text says "AWS Lambda" specifically

3. NEVER pick from a completely different domain.
   ✗ "Compliance Strategies" → "Digital Content Strategy" — completely wrong domain
   If no candidate fits, set no_match: true

4. PREFER the most specific correct match over a generic one.
   "Sorting Algorithms" → prefer "Sorting Algorithm" over "Algorithms"
   Use raw_skill_keywords to disambiguate when candidates are close.

════════════════════
OUTPUT — EXACTLY 1 MATCH PER SKILL
════════════════════
Return the lightcast_id and match_confidence ONLY.
Do NOT return mckinsey fields — those will be looked up from the database.

Skills to map:
{skills_block}

Return ONLY a valid JSON array — one object per skill, same order as input:
[
  {{
    "skill_index": 0,
    "lightcast_id": "<exact id from the chosen candidate row, or empty if no_match>",
    "match_confidence": 0.0,
    "no_match": false,
    "no_match_reason": ""
  }}
]
No markdown fences. No explanation. Only the JSON array."""


def _gemini_map_lightcast_batch(skills_payload: list) -> list:
    """
    Send a batch of skills with their top-15 LightCast candidates to Gemini.
    Gemini returns ONE lightcast_id per skill (+ match_confidence + no_match).
    mckinsey fields are NOT sent to Gemini — they are looked up from the
    candidate dict after Gemini picks the id.
    """
    skills_block = ""
    for item in skills_payload:
        skills_block += f"\n--- Skill index {item['index']} ---\n"
        skills_block += f"Skill name         : {item['skill']}\n"
        if item.get("raw_skill_keywords"):
            skills_block += f"raw_skill_keywords : {', '.join(item['raw_skill_keywords'])}\n"
        if item.get("subject"):
            skills_block += f"From subject       : {item['subject']}\n"
        if item.get("bloom_level"):
            skills_block += f"Bloom\'s level      : {item['bloom_level']}\n"
        if item.get("bloom_explanation"):
            skills_block += f"Bloom\'s explanation: {item['bloom_explanation']}\n"
        skills_block += "Candidates (rank | id | name | category | similarity):\n"
        for c in item.get("candidates", []):
            skills_block += (
                f"  {c['rank']:>2}. id={c['lightcast_id']:<36} "
                f"| {c['talent_skill']:<45} "
                f"| {c['mckinsey_category']:<22} "
                f"| sim={c['similarity']}\n"
            )

    prompt = LIGHTCAST_BATCH_PROMPT.format(skills_block=skills_block)
    result = _call_gemini_json(prompt, max_tokens=2048)
    return result if isinstance(result, list) else []


def map_skills_to_lightcast(rows: list[dict], emit) -> list[dict]:
    """
    Map every extracted skill to EXACTLY ONE LightCast skill.

    Flow:
      1. OpenAI embed all unique skill names (one batch call)
      2. Cosine search → top-15 candidates per skill from in-memory LightCast cache
      3. Gemini picks the single best lightcast_id per skill (nothing else)
      4. All fields (lightcast_skill, mckinsey_*) resolved from the candidate dict
         — NOT from Gemini output, which prevents "None" hallucinations
      5. Dedup rows on (subject, skill)

    row["skill"] is never overwritten — stays as the Gemini-extracted name.
    """
    def _mark_unmatched(rows):
        for row in rows:
            row.setdefault("lightcast_skills", [])
            row.setdefault("lightcast_source", "unmatched")
        return rows

    if not rows:
        return rows

    if not OPENAI_API_KEY:
        emit("progress", {"step": "lightcast",
            "msg": "OPENAI_API_KEY not set — LightCast mapping skipped.", "pct": 92})
        return _mark_unmatched(rows)

    lc_docs = _get_lightcast_docs()
    if not lc_docs:
        emit("progress", {"step": "lightcast",
            "msg": "LightCast collection empty — mapping skipped.", "pct": 92})
        return _mark_unmatched(rows)

    total_skills = len(rows)
    emit("progress", {"step": "lightcast",
        "msg": f"Mapping {total_skills} skills → LightCast taxonomy (embedding…)", "pct": 86})

    # ── Step 1: Embed unique skill names ──────────────────────────────────────
    unique_skills = list(dict.fromkeys(r.get("skill", "") for r in rows))
    try:
        embeddings = generate_openai_embeddings(unique_skills)
        emb_map    = dict(zip(unique_skills, embeddings))
    except Exception as e:
        emit("progress", {"step": "lightcast",
            "msg": f"Embedding failed ({e}) — skipped.", "pct": 92})
        return _mark_unmatched(rows)

    # ── Step 2: Cosine top-15 per skill ───────────────────────────────────────
    skill_candidates: dict[str, list] = {}
    for skill in unique_skills:
        emb = emb_map.get(skill, [])
        skill_candidates[skill] = _cosine_top_k(emb, lc_docs, top_k=15) if emb else []

    # Build id → full candidate dict for instant lookup after Gemini picks
    # { skill_name: { lightcast_id: candidate_dict } }
    cand_by_id: dict[str, dict[str, dict]] = {
        skill: {c["lightcast_id"]: c for c in cands}
        for skill, cands in skill_candidates.items()
    }

    # ── Step 3: Gemini picks one lightcast_id per skill ───────────────────────
    BATCH_SIZE = 8
    row_ctx: dict[str, dict] = {}
    for r in rows:
        s = r.get("skill", "")
        if s not in row_ctx:
            row_ctx[s] = r

    unique_with_cands = [(s, skill_candidates[s]) for s in unique_skills if skill_candidates.get(s)]
    lightcast_map: dict[str, dict] = {}
    total_batches = max(1, (len(unique_with_cands) + BATCH_SIZE - 1) // BATCH_SIZE)

    for batch_start in range(0, len(unique_with_cands), BATCH_SIZE):
        batch     = unique_with_cands[batch_start: batch_start + BATCH_SIZE]
        batch_num = batch_start // BATCH_SIZE + 1

        payload = []
        for i, (skill, cands) in enumerate(batch):
            ctx = row_ctx.get(skill, {})
            payload.append({
                "index":             i,
                "skill":             skill,
                "raw_skill_keywords": ctx.get("raw_skill_keywords", []),
                "subject":           ctx.get("subject",     ""),
                "bloom_level":       ctx.get("bloom_level", ""),
                "bloom_explanation": ctx.get("bloom_explanation", ""),
                "candidates": [
                    {
                        "rank":              j + 1,
                        "lightcast_id":      c["lightcast_id"],
                        "talent_skill":      c["talent_skill"],
                        "mckinsey_category": c["mckinsey_category"],
                        "similarity":        c["similarity"],
                    }
                    for j, c in enumerate(cands)
                ],
            })

        try:
            results = _gemini_map_lightcast_batch(payload)
            for res in results:
                idx = res.get("skill_index", res.get("index", -1))
                if isinstance(idx, int) and 0 <= idx < len(batch):
                    lightcast_map[batch[idx][0]] = res
        except Exception as e:
            emit("progress", {"step": "lightcast",
                "msg": f"Batch {batch_num}/{total_batches} error: {e} — continuing."})

        pct = 86 + int((batch_start + BATCH_SIZE) / max(len(unique_with_cands), 1) * 6)
        emit("progress", {"step": "lightcast",
            "msg": f"LightCast: batch {batch_num}/{total_batches} mapped.", "pct": min(pct, 92)})

    # ── Step 4: Resolve all fields from candidate dict — NOT from Gemini ──────
    matched = 0
    for row in rows:
        gemini_skill = row.get("skill", "")
        gemini_res   = lightcast_map.get(gemini_skill, {})

        if gemini_res and not gemini_res.get("no_match"):
            picked_id  = str(gemini_res.get("lightcast_id", "")).strip()
            confidence = float(gemini_res.get("match_confidence", 0.0))

            # Look up ALL fields from our own candidate data (guaranteed correct)
            cand = cand_by_id.get(gemini_skill, {}).get(picked_id)
            if not cand:
                # Gemini returned an unrecognised id — fall back to cosine top-1
                top = skill_candidates.get(gemini_skill, [])
                cand = top[0] if top else None
                if cand:
                    confidence = cand["similarity"]

            if cand:
                row["lightcast_skills"] = [{
                    "lightcast_id":     cand["lightcast_id"],
                    "lightcast_skill":  cand["talent_skill"],
                    "match_confidence": confidence,
                }]
                row["lightcast_source"] = "verified"
                matched += 1
            else:
                row["lightcast_skills"] = []
                row["lightcast_source"] = "unmatched"
        else:
            row["lightcast_skills"] = []
            row["lightcast_source"] = "unmatched"
            if gemini_res and gemini_res.get("no_match_reason"):
                row["lightcast_no_match_reason"] = gemini_res["no_match_reason"]

    # ── Step 5: Dedup rows on (subject, skill) ────────────────────────────────
    seen: set = set()
    deduped: list[dict] = []
    for row in rows:
        key = (row.get("subject", "").strip().lower(), row.get("skill", "").strip().lower())
        if key not in seen:
            seen.add(key)
            deduped.append(row)

    removed = len(rows) - len(deduped)
    emit("progress", {"step": "lightcast",
        "msg": (
            f"LightCast: {matched}/{total_skills} skills mapped. "
            + (f"{total_skills - matched} unmatched. " if total_skills - matched else "")
            + (f"{removed} duplicates removed." if removed else "")
        ), "pct": 93})
    return deduped


# ══════════════════════════════════════════════════════════════════════════════
# JSON helpers (salvage_json unchanged — copy from original)
# ══════════════════════════════════════════════════════════════════════════════

def salvage_json(raw: str) -> list:
    raw = raw.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.MULTILINE)
    raw = re.sub(r"\s*```\s*$",       "", raw, flags=re.MULTILINE)
    raw = raw.strip()
    start = raw.find("[")
    if start == -1:
        return []
    end = raw.rfind("]")
    if end != -1 and end > start:
        try:
            return json.loads(raw[start: end + 1])
        except json.JSONDecodeError:
            pass
    chunk = raw[start:]
    last_brace = chunk.rfind("}")
    if last_brace == -1:
        return []
    salvaged = chunk[: last_brace + 1].rstrip().rstrip(",") + "\n]"
    try:
        return json.loads(salvaged)
    except json.JSONDecodeError:
        pass
    objects = re.findall(r'\{[^{}]+\}', chunk, re.DOTALL)
    result  = []
    for obj_str in objects:
        try:
            result.append(json.loads(obj_str))
        except json.JSONDecodeError:
            continue
    return result


# ══════════════════════════════════════════════════════════════════════════════
# GEMINI EXTRACTION (call_gemini_chunk, deduplicate, extract_skills_gemini)
# ══════════════════════════════════════════════════════════════════════════════

def call_gemini_chunk(client, chunk_text, department, year, semester, chunk_idx, total_chunks):
    prompt = EXTRACTION_SYSTEM_PROMPT + "\n\n" + EXTRACTION_PROMPT.format(
        department=department, year=year, semester=semester,
        text=chunk_text, chunk_idx=chunk_idx, total_chunks=total_chunks,
    )
    def _call():
        response = client.models.generate_content(
            model=GEMINI_MODEL,
            contents=prompt,
            config=genai_types.GenerateContentConfig(
                temperature=0.1,
                max_output_tokens=65_536,
            ),
        )
        return salvage_json(response.text or "")
    return _gemini_with_retry(_call)


def deduplicate(rows: list[dict]) -> list[dict]:
    seen = set()
    out  = []
    for row in rows:
        key = (row.get("subject", "").strip().lower(),
               row.get("skill",   "").strip().lower())
        if key not in seen:
            seen.add(key)
            out.append(row)
    return out


# ══════════════════════════════════════════════════════════════════════════════
# MCKINSEY SKILL CLASSIFICATION
# ══════════════════════════════════════════════════════════════════════════════

MCKINSEY_TAXONOMY = {
    "Cognitive": {
        "Critical Thinking": [
            "Structured problem solving", "Logical reasoning",
            "Understanding biases", "Seeking relevant information",
        ],
        "Planning and Ways of Working": [
            "Work-plan development", "Time management and prioritization", "Agile thinking",
        ],
        "Communication": [
            "Storytelling and public speaking", "Asking the right questions",
            "Synthesizing messages", "Active listening",
        ],
        "Mental Flexibility": [
            "Creativity and imagination", "Translating knowledge to different contexts",
            "Adopting a different perspective", "Adaptability", "Ability to learn",
        ],
    },
    "Interpersonal": {
        "Mobilizing Systems": [
            "Role modeling", "Win-win negotiations",
            "Crafting an inspiring vision", "Organizational awareness",
        ],
        "Developing Relationships": [
            "Empathy", "Inspiring trust", "Humility", "Sociability",
        ],
        "Teamwork Effectiveness": [
            "Fostering inclusiveness", "Motivating different personalities",
            "Resolving conflicts", "Collaboration", "Coaching", "Empowering",
        ],
    },
    "Self-Leadership": {
        "Self-Awareness and Self-Management": [
            "Understanding own emotions and triggers", "Self-control and regulation",
            "Understanding own strengths", "Integrity",
            "Self-motivation and wellness", "Self-confidence",
        ],
        "Entrepreneurship": [
            "Courage and risk-taking", "Driving change and innovation",
            "Energy, passion, and optimism", "Breaking orthodoxies",
        ],
        "Goals Achievement": [
            "Ownership and decisiveness", "Achievement orientation",
            "Grit and persistence", "Coping with uncertainty", "Self-development",
        ],
    },
    "Digital": {
        "Digital Fluency and Citizenship": [
            "Digital literacy", "Digital learning", "Digital collaboration", "Digital ethics",
        ],
        "Software Use and Development": [
            "Programming literacy", "Data analysis and statistics",
            "Computational and algorithmic thinking",
        ],
        "Understanding Digital Systems": [
            "Data literacy", "Smart systems",
            "Cybersecurity literacy", "Tech translation and enablement",
        ],
    },
}

_MCKINSEY_FLAT = "\n".join(
    f"  Category: {cat} | Subcategory: {sub} | Sub-subcategory: {ss}"
    for cat, subs in MCKINSEY_TAXONOMY.items()
    for sub, sss in subs.items()
    for ss in sss
)

MCKINSEY_PROMPT = """You are a skills taxonomy expert. Classify each curriculum skill
into the McKinsey Skills Framework.

FRAMEWORK (use ONLY values from this list — copy them exactly):
{taxonomy}

RULES:
1. Choose based on the SKILL meaning, not the subject name.
2. All three levels must come from the framework above.
3. Most technical/programming/algorithm/cloud skills -> Digital > Software Use and Development.
4. Skills about people/organisations/leadership -> Interpersonal or Self-Leadership.
5. If a skill spans two categories, pick the PRIMARY one.
6. NEVER invent values. Copy exact strings from the framework.

Skills to classify:
{skills_block}

Return ONLY a valid JSON array — one object per skill, same order:
[
  {{
    "skill_index": 0,
    "mckinsey_category":        "<exact Category string>",
    "mckinsey_subcategory":     "<exact Subcategory string>",
    "mckinsey_sub_subcategory": "<exact Sub-subcategory string>"
  }}
]
No markdown fences. No explanation. Only the JSON array."""


def classify_skills_mckinsey(rows: list[dict], emit) -> list[dict]:
    """
    Batch-classify every Gemini-extracted skill into the McKinsey framework.
    Sets top-level row fields: mckinsey_category, mckinsey_subcategory, mckinsey_sub_subcategory.
    """
    if not rows:
        return rows

    unique_keys = list(dict.fromkeys(
        (r.get("skill", ""), r.get("subject", "")) for r in rows
    ))

    emit("progress", {"step": "mckinsey",
        "msg": f"Classifying {len(unique_keys)} skills into McKinsey framework...", "pct": 94})

    BATCH_SIZE = 12
    mckinsey_map: dict = {}

    for batch_start in range(0, len(unique_keys), BATCH_SIZE):
        batch = unique_keys[batch_start: batch_start + BATCH_SIZE]

        skills_block = ""
        for i, (skill, subject) in enumerate(batch):
            skills_block += f"\n  {i}. Skill: \"{skill}\""
            if subject:
                skills_block += f" (from: {subject})"

        prompt = MCKINSEY_PROMPT.format(taxonomy=_MCKINSEY_FLAT, skills_block=skills_block)

        try:
            results = _call_gemini_json(prompt, max_tokens=2048)
            if isinstance(results, list):
                for res in results:
                    idx = res.get("skill_index", -1)
                    if isinstance(idx, int) and 0 <= idx < len(batch):
                        mckinsey_map[batch[idx]] = {
                            "mckinsey_category":        res.get("mckinsey_category",        ""),
                            "mckinsey_subcategory":     res.get("mckinsey_subcategory",     ""),
                            "mckinsey_sub_subcategory": res.get("mckinsey_sub_subcategory", ""),
                        }
        except Exception as e:
            emit("progress", {"step": "mckinsey", "msg": f"McKinsey batch error: {e} -- continuing."})

    valid_cats    = set(MCKINSEY_TAXONOMY.keys())
    valid_subs    = {s for subs in MCKINSEY_TAXONOMY.values() for s in subs}
    valid_subsubs = {ss for subs in MCKINSEY_TAXONOMY.values()
                     for sss in subs.values() for ss in sss}

    for row in rows:
        key = (row.get("skill", ""), row.get("subject", ""))
        mc  = mckinsey_map.get(key, {})
        cat    = mc.get("mckinsey_category",        "")
        sub    = mc.get("mckinsey_subcategory",     "")
        subsub = mc.get("mckinsey_sub_subcategory", "")
        row["mckinsey_category"]        = cat    if cat    in valid_cats    else ""
        row["mckinsey_subcategory"]     = sub    if sub    in valid_subs    else ""
        row["mckinsey_sub_subcategory"] = subsub if subsub in valid_subsubs else ""

    classified = sum(1 for r in rows if r.get("mckinsey_category"))
    emit("progress", {"step": "mckinsey",
        "msg": f"McKinsey: {classified}/{len(rows)} skills classified.", "pct": 96})
    return rows


def extract_skills_gemini(text, department, year, semester, emit) -> list[dict]:
    client = get_genai_client()
    chunks = split_text_into_chunks(text)
    total  = len(chunks)

    emit("progress", {
        "step": "ai",
        "msg": f"Sending {total} chunk(s) ({len(text):,} chars) to Gemini…",
        "pct": 45,
    })

    results_by_idx = [[] for _ in range(total)]
    lock = Lock()
    done = [0]

    def process(idx, chunk):
        rows = call_gemini_chunk(client, chunk, department, year, semester, idx + 1, total)
        with lock:
            done[0] += 1
            pct = 45 + int(done[0] / total * 35)
            emit("progress", {
                "step": "ai",
                "msg": f"Gemini: chunk {done[0]}/{total} → {len(rows)} skills",
                "pct": pct,
            })
        return idx, rows

    futs = [_executor.submit(process, i, c) for i, c in enumerate(chunks)]
    for fut in as_completed(futs):
        idx, rows = fut.result()
        results_by_idx[idx] = rows

    merged = []
    for rows in results_by_idx:
        merged.extend(rows)

    deduped = deduplicate(merged)
    emit("progress", {
        "step": "ai",
        "msg": f"Gemini: {len(deduped)} unique skills. Mapping to LightCast…",
        "pct": 82,
    })

    deduped = map_skills_to_lightcast(deduped, emit)

    # McKinsey classification — uses Gemini skill name, independent of LightCast
    deduped = classify_skills_mckinsey(deduped, emit)
    return deduped


# ══════════════════════════════════════════════════════════════════════════════
# ★ UPDATED EXCEL BUILDER — handles lightcast_skills list + raw_skill_keywords
# ══════════════════════════════════════════════════════════════════════════════

HEADER_BG = "1E293B"
HEADER_FG = "FFFFFF"
ALT_BG    = "F8FAFC"
NORMAL_BG = "FFFFFF"

LC_VERIFIED_BG  = "DCFCE7"
LC_VERIFIED_FG  = "166534"
LC_UNMATCHED_BG = "FEF9C3"
LC_UNMATCHED_FG = "713F12"

PROF_STYLE = {
    "Beginner":     {"bg": "DBEAFE", "fg": "1E40AF"},
    "Intermediate": {"bg": "D1FAE5", "fg": "065F46"},
    "Advanced":     {"bg": "F3E8FF", "fg": "6B21A8"},
}
BLOOM_STYLE = {
    "Remember":   {"bg": "FEE2E2", "fg": "991B1B"},
    "Understand": {"bg": "FEF3C7", "fg": "92400E"},
    "Apply":      {"bg": "D1FAE5", "fg": "065F46"},
    "Analyze":    {"bg": "DBEAFE", "fg": "1E40AF"},
    "Evaluate":   {"bg": "EDE9FE", "fg": "5B21B6"},
    "Create":     {"bg": "FCE7F3", "fg": "9D174D"},
    "Inferred":   {"bg": "F3F4F6", "fg": "374151"},
}


def thin_border(color="D1D5DB"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def create_curriculum_excel(data: list[dict], department: str, year: str, semester: str) -> io.BytesIO:
    """
    Build the curriculum skill Excel.

    Column changes vs original:
      • Col L  : LightCast Skill(s)  — primary name + secondary names in grey below (same cell, newlines)
      • Col M  : Raw Skill Keywords   — new column showing specific tools/APIs named in unit
      • Col N  : LC Category
      • Col O  : Proficiency
      • (cols shifted accordingly)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MCA Semester 3"

    headers = [
        "Department", "Year", "Semester", "Subject", "Credits", "Hours",
        "Bloom's Level", "Bloom's Explanation", "Skill (Gemini)", "Raw Skill Keywords",
        "McKinsey Category", "McKinsey Subcategory", "McKinsey Sub-subcategory",
        "LightCast Skill", "Match Confidence", "Proficiency", "Proficiency Rationale",
    ]

    col_widths = [14, 12, 12, 36, 9, 8, 14, 45, 36, 36, 22, 26, 32, 36, 14, 14, 60]

    for ci, (hdr, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.fill      = PatternFill("solid", fgColor=HEADER_BG)
        c.font      = Font(bold=True, color=HEADER_FG, size=11, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border("1E293B")
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 32

    for ri, item in enumerate(data, 2):
        row_bg = ALT_BG if ri % 2 == 0 else NORMAL_BG
        bloom  = str(item.get("bloom_level", "")).strip()
        prof   = str(item.get("proficiency", "")).strip()
        lc_src = item.get("lightcast_source", "")

        # LightCast
        lc_list      = item.get("lightcast_skills", [])
        lc_name      = lc_list[0].get("lightcast_skill", "") if lc_list else ""
        lc_conf      = lc_list[0].get("match_confidence", 0) if lc_list else 0

        # McKinsey — top-level row fields set by classify_skills_mckinsey
        mck_cat    = item.get("mckinsey_category",        "")
        mck_sub    = item.get("mckinsey_subcategory",     "")
        mck_subsub = item.get("mckinsey_sub_subcategory", "")

        # Raw keywords
        kws        = item.get("raw_skill_keywords", [])
        kw_display = ", ".join(kws) if kws else ""

        vals = [
            item.get("department",   department),
            item.get("year",         year),
            item.get("semester",     semester),
            item.get("subject",      ""),
            item.get("credits",      ""),
            item.get("total_hours",  ""),
            bloom,
            item.get("bloom_explanation", ""),  # Bloom's explanation
            item.get("skill",        ""),     # Gemini skill name
            kw_display,                        # Raw keywords
            mck_cat,                           # McKinsey Category
            mck_sub,                           # McKinsey Subcategory
            mck_subsub,                        # McKinsey Sub-subcategory
            lc_name,                           # LightCast Skill
            f"{lc_conf:.3f}" if lc_conf else "",
            prof,
            item.get("proficiency_rationale", ""),
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border    = thin_border()
            c.font      = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="center", wrap_text=(ci in (4, 8, 9, 10, 14, 17)))

            if ci == 7:    # Bloom's Level
                style = BLOOM_STYLE.get(bloom, BLOOM_STYLE["Inferred"])
                c.fill = PatternFill("solid", fgColor=style["bg"])
                c.font = Font(bold=True, color=style["fg"], name="Calibri", size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 14:  # LightCast Skill
                if lc_src == "verified":
                    c.fill = PatternFill("solid", fgColor=LC_VERIFIED_BG)
                    c.font = Font(bold=True, color=LC_VERIFIED_FG, name="Calibri", size=10)
                else:
                    c.fill = PatternFill("solid", fgColor=LC_UNMATCHED_BG)
                    c.font = Font(italic=True, name="Calibri", size=10, color=LC_UNMATCHED_FG)
            elif ci == 16:  # Proficiency
                style = PROF_STYLE.get(prof, {})
                if style:
                    c.fill = PatternFill("solid", fgColor=style["bg"])
                    c.font = Font(bold=True, color=style["fg"], name="Calibri", size=10)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.fill = PatternFill("solid", fgColor=row_bg)
            else:
                c.fill = PatternFill("solid", fgColor=row_bg)

        ws.row_dimensions[ri].height = 30

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P{len(data)+1}"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 24

    ws2["A1"] = "Curriculum Skill Extraction Summary"
    ws2["A1"].font = Font(bold=True, size=14, name="Calibri")

    lc_verified  = sum(1 for d in data if d.get("lightcast_source") == "verified")
    lc_unmatched = sum(1 for d in data if d.get("lightcast_source") == "unmatched")
    mck_classified = sum(1 for d in data if d.get("mckinsey_category"))

    for r, (k, vv) in enumerate([
        ("Department", department), ("Year", year), ("Semester", semester),
        ("Total Skills", len(data)),
        ("Unique Subjects", len({d.get("subject", "") for d in data})),
        ("LightCast Verified Skills", lc_verified),
        ("Unmatched Skills (original kept)", lc_unmatched),
        ("McKinsey Classified Skills", mck_classified),
    ], 3):
        ws2.cell(row=r, column=1, value=k).font = Font(bold=True, name="Calibri")
        ws2.cell(row=r, column=2, value=vv)

    ws2["A13"] = "McKinsey Category Distribution"
    ws2["A13"].font = Font(bold=True, name="Calibri", size=12)
    for r, (cat, n) in enumerate(Counter(d.get("mckinsey_category", "") for d in data if d.get("mckinsey_category")).items(), 14):
        ws2.cell(row=r, column=1, value=cat)
        ws2.cell(row=r, column=2, value=n)

    ws2["A20"] = "Proficiency Breakdown"
    ws2["A20"].font = Font(bold=True, name="Calibri", size=12)
    for r, (p, n) in enumerate(Counter(d.get("proficiency", "") for d in data).items(), 21):
        c1 = ws2.cell(row=r, column=1, value=p)
        ws2.cell(row=r, column=2, value=n)
        style = PROF_STYLE.get(p, {})
        if style:
            c1.fill = PatternFill("solid", fgColor=style["bg"])
            c1.font = Font(bold=True, color=style["fg"], name="Calibri")

    ws2["A28"] = "Bloom's Level Distribution"
    ws2["A28"].font = Font(bold=True, name="Calibri", size=12)
    for r, (bl, n) in enumerate(Counter(d.get("bloom_level", "Inferred") for d in data).items(), 29):
        c1 = ws2.cell(row=r, column=1, value=bl)
        ws2.cell(row=r, column=2, value=n)
        style = BLOOM_STYLE.get(bl, BLOOM_STYLE["Inferred"])
        c1.fill = PatternFill("solid", fgColor=style["bg"])
        c1.font = Font(bold=True, color=style["fg"], name="Calibri")

    ws2["A36"] = "Skills per Subject"
    ws2["A36"].font = Font(bold=True, name="Calibri", size=12)
    for r, (s, n) in enumerate(Counter(d.get("subject", "") for d in data).items(), 37):
        ws2.cell(row=r, column=1, value=s)
        ws2.cell(row=r, column=2, value=n)

    # ── CO Evidence sheet ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("CO Evidence")
    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 36
    ws3.column_dimensions["C"].width = 30
    ws3.column_dimensions["D"].width = 22
    ws3.column_dimensions["E"].width = 26
    ws3.column_dimensions["F"].width = 32
    ws3.column_dimensions["G"].width = 36
    ws3.column_dimensions["H"].width = 16
    ws3.column_dimensions["I"].width = 45
    ws3.column_dimensions["J"].width = 14
    ws3.column_dimensions["K"].width = 60

    co_hdrs = [
        "Subject", "Skill (Gemini)", "Raw Keywords",
        "McKinsey Category", "McKinsey Subcategory", "McKinsey Sub-subcategory",
        "LightCast Skill", "Bloom's Level", "Bloom's Explanation", "Proficiency", "Rationale",
    ]
    for ci, hdr in enumerate(co_hdrs, 1):
        c = ws3.cell(row=1, column=ci, value=hdr)
        c.fill      = PatternFill("solid", fgColor=HEADER_BG)
        c.font      = Font(bold=True, color=HEADER_FG, size=11, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border("1E293B")
    ws3.row_dimensions[1].height = 26

    for ri, item in enumerate(data, 2):
        row_bg = ALT_BG if ri % 2 == 0 else NORMAL_BG
        bloom  = str(item.get("bloom_level", "")).strip()
        prof   = str(item.get("proficiency", "")).strip()
        lc_src = item.get("lightcast_source", "")

        lc_list    = item.get("lightcast_skills", [])
        lc_name    = lc_list[0].get("lightcast_skill", "") if lc_list else ""
        kws        = item.get("raw_skill_keywords", [])
        kw_display = ", ".join(kws) if kws else ""

        vals = [
            item.get("subject",               ""),
            item.get("skill",                  ""),
            kw_display,
            item.get("mckinsey_category",        ""),
            item.get("mckinsey_subcategory",     ""),
            item.get("mckinsey_sub_subcategory", ""),
            lc_name,
            bloom,
            item.get("bloom_explanation", ""),
            prof,
            item.get("proficiency_rationale", ""),
        ]
        for ci, val in enumerate(vals, 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.border    = thin_border()
            c.font      = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="center", wrap_text=(ci in (1, 2, 3, 9, 11)))

            if ci == 7:    # LightCast Skill
                if lc_src == "verified":
                    c.fill = PatternFill("solid", fgColor=LC_VERIFIED_BG)
                    c.font = Font(bold=True, color=LC_VERIFIED_FG, name="Calibri", size=10)
                else:
                    c.fill = PatternFill("solid", fgColor=LC_UNMATCHED_BG)
                    c.font = Font(italic=True, name="Calibri", size=10, color=LC_UNMATCHED_FG)
            elif ci == 8:  # Bloom's Level
                style = BLOOM_STYLE.get(bloom, BLOOM_STYLE["Inferred"])
                c.fill = PatternFill("solid", fgColor=style["bg"])
                c.font = Font(bold=True, color=style["fg"], name="Calibri", size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 10:  # Proficiency
                style = PROF_STYLE.get(prof, {})
                if style:
                    c.fill = PatternFill("solid", fgColor=style["bg"])
                    c.font = Font(bold=True, color=style["fg"], name="Calibri", size=10)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.fill = PatternFill("solid", fgColor=row_bg)
            else:
                c.fill = PatternFill("solid", fgColor=row_bg)
        ws3.row_dimensions[ri].height = 30

    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:J{len(data)+1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
# SSE BACKGROUND JOB RUNNER
# ══════════════════════════════════════════════════════════════════════════════

def run_extraction_job(job_id, pdf_bytes, department, year, semester, save_to_mongo: bool):
    q = _job_store.get(job_id)
    if q is None:
        return

    def emit(event, data):
        q.put(sse_msg(event, data))

    try:
        emit("progress", {"step": "upload", "msg": "PDF received. Starting extraction…", "pct": 5})

        text, method = extract_full_text(pdf_bytes, emit)
        char_count   = len(text)
        n_chunks     = len(split_text_into_chunks(text))

        emit("progress", {"step": "extract",
            "msg": f"Extracted {char_count:,} chars via {method}. {n_chunks} Gemini chunk(s).",
            "pct": 42})

        if char_count < 100:
            emit("error", {"msg": "Could not extract readable text from the PDF."})
            return

        data = extract_skills_gemini(text, department, year, semester, emit)

        if not data:
            emit("error", {"msg": "Gemini returned no skills. PDF may be image-only."})
            return

        mongo_result = {}
        if save_to_mongo:
            emit("progress", {"step": "mongo", "msg": "Saving LightCast-mapped skills to MongoDB…", "pct": 94})
            replace_filter = {"department": department, "semester": semester}
            mongo_result = save_extracted_to_mongo(data, replace_filter=replace_filter)
            emit("progress", {"step": "mongo",
                "msg": f"MongoDB: {mongo_result.get('inserted', 0)} inserted, {mongo_result.get('modified', 0)} updated.",
                "pct": 96})

        emit("progress", {"step": "excel", "msg": f"Building Excel ({len(data)} rows)…", "pct": 97})
        excel_buf = create_curriculum_excel(data, department, year, semester)

        with _result_lock:
            _result_store.put(job_id, excel_buf.getvalue())

        lc_verified = sum(1 for d in data if d.get("lightcast_source") == "verified")
        emit("progress", {"step": "excel", "msg": "Excel ready.", "pct": 100})

        # Remove embedding fields from preview data to reduce response size
        preview_data = []
        for item in data[:150]:
            cleaned_item = {k: v for k, v in item.items()
                          if k not in ('skill_embedding', 'description_embedding')}
            preview_data.append(cleaned_item)

        emit("done", {
            "job_id":       job_id,
            "count":        len(data),
            "lc_verified":  lc_verified,
            "lc_unmatched": len(data) - lc_verified,
            "method":       method,
            "subjects":     len({d.get("subject", "") for d in data}),
            "chars":        char_count,
            "chunks":       n_chunks,
            "mongo":        mongo_result,
            "preview":      preview_data,
        })

    except Exception as e:
        emit("error", {"msg": str(e), "trace": traceback.format_exc()})
    finally:
        q.put(None)