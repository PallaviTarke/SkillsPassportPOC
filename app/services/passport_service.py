"""
Passport generation service - handles marksheet parsing and skill passport assembly.
"""

from __future__ import annotations

import io
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from typing import Optional

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import openpyxl

from app.core.config import _call_gemini_json, log
from app.core.database import get_curriculum
from app.services.extraction_service import (
    extract_full_text, PROF_STYLE, HEADER_BG, HEADER_FG,
    ALT_BG, NORMAL_BG, thin_border
)

# ══════════════════════════════════════════════════════════════════════════════
# PROMPTS
# ══════════════════════════════════════════════════════════════════════════════

MARKSHEET_PARSE_PROMPT = """You are parsing a student academic result / marksheet document.

YOUR ONLY JOB: extract structured data from the text. Do NOT infer skills. Do NOT guess.

━━━ STEP 0 — DETECT THE UNIVERSITY'S GRADE SCHEME FIRST ━━━

Before reading any subject rows, scan the ENTIRE document for a grade/point table.
Universities often print their own grade scale. Examples:

  DYPIU format:   A+=10, A=9, B+=8, B=7, C+=6, C=5, D=4, F=0
  SPPU format:    O=10, A+=9, A=8, B+=7, B=6, C=5, P=4, F=0
  Standard 10pt:  O=10, A+=9, A=8, B+=7, B=6, C=5, D=4, P=4, F=0
  Raw marks:      no grade table — marks_obtained / marks_total given directly

If you find a grade table in the document → use THOSE exact grade-to-point values.
If no grade table found → use this fallback: O=10, A+=9, A=8, B+=7, B=6, C=5, P=4, F=0

Record the detected scheme as "grade_scheme" in your output.

━━━ WHAT TO EXTRACT ━━━

1. STUDENT INFO
   name, roll_no (seat / PRN / enrollment), department (branch/course),
   semester, year (academic year), college, sgpa (float or null)

2. EVERY SUBJECT ROW
   subject_name  — clean name WITHOUT the subject code
   sub_code      — subject code if present (e.g. CSE1101), else null
   grade         — letter grade exactly as printed (F / B / A+ / D / N/A etc.)
   gp            — grade point from the detected scheme above (null if N/A or not applicable)
   credits       — credit value as float (e.g. 2.5), else null
   marks_obtained — float if raw marks available, else null
   marks_total    — float if raw marks available, else null
   score_percentage — calculate using THIS priority:
       1. raw marks available → (marks_obtained / marks_total) × 100
       2. gp available        → (gp / max_gp) × 100   where max_gp = highest point in detected scheme
       3. grade only          → convert using detected scheme, then (points / max_gp) × 100
       4. grade is N/A, I, R, or Incomplete → null
       5. nothing found       → null

   IMPORTANT: For F grade → gp = 0 → score_percentage = 0.0 (not null)
   IMPORTANT: For N/A, Rural Internship, Audit courses → score_percentage = null

3. RESULT FORMAT
   "raw_marks" | "grade_gp" | "grade_only"

4. OVERALL PERCENTAGE  (SGPA × 10, or computed, or null)

━━━ IMPORTANT ━━━
• subject_name must be the CLEAN NAME only — strip any code prefix.
• Include ALL subjects including failed ones (F grade).
• Rural Internship / audit / N/A subjects — include them but set score_percentage = null.
• Do not invent subjects not in the text.

Return ONLY valid JSON, no markdown fences:
{
  "student_info": {
    "name": "...", "roll_no": "...", "department": "...",
    "semester": "...", "year": "...", "college": "...", "sgpa": null
  },
  "grade_scheme": "DYPIU: A+=10, A=9, B+=8, B=7, C+=6, C=5, D=4, F=0",
  "result_format": "grade_gp",
  "overall_percentage": null,
  "subjects_extracted": [
    {
      "subject_name": "Data Structures",
      "sub_code": "CSE1101",
      "grade": "F", "gp": 0, "credits": 5,
      "marks_obtained": null, "marks_total": null,
      "score_percentage": 0.0
    }
  ]
}"""


SKILLS_INFER_PROMPT = """You are an industry skills taxonomist.
A student's marksheet has been parsed. No curriculum database is available.
Infer exactly 2–3 hybrid skills per subject — noun phrases (2–5 words), industry-recognisable.
Do NOT output more than 3 skills per subject. 2 is fine for narrow subjects, 3 for broad ones.

━━━ CONTEXT: NO CURRICULUM DATABASE AVAILABLE ━━━

This subject was NOT found in the curriculum database. You only have the subject name,
subject code, and credit value from the marksheet — no syllabus content exists.
Because there is no curriculum text to analyse, credits are the BEST available proxy
for how deeply the subject was taught. Use them as the primary signal.

━━━ CRITICAL: PROFICIENCY = CURRICULUM DEPTH, NOT STUDENT SCORE ━━━

inferred_proficiency describes how DEEP the syllabus likely covers the skill.
It has NOTHING to do with whether the student passed or failed.

DECISION ORDER — work through these in order, stop when you have enough evidence:

  SIGNAL 1 — Subject name vocabulary (use first):
    "Introduction to X" / "Basics of X" / "Fundamentals of X" → Beginner
    "Applied X" / "X Design" / "Analysis of X" / "X Engineering" → Intermediate
    "Advanced X" / "X Optimisation" / "X Architecture" / capstone → Advanced
    Workshop / Lab / Internship / Seminar → Beginner (exposure only)
    If the name is neutral (e.g. "Data Structures", "Networks") → move to Signal 2.

  SIGNAL 2 — Credits (primary fallback when subject name is neutral):
    No curriculum text exists, so credits are the best depth indicator available.
    1–2 credits → Beginner
    3–4 credits → Intermediate
    5+  credits → Intermediate or Advanced (never Beginner for 5+ credit core subjects)
    A 2-credit workshop/lab is always Beginner regardless of semester.

  SIGNAL 3 — Semester / year position (corroboration only):
    Semester 1–2 → lean Beginner
    Semester 3–5 → lean Intermediate
    Semester 6–8 → lean Intermediate / Advanced
    Use this to break ties between Signals 1 and 2, not to override them.

  DECISION RULE:
    Subject name is the clearest signal when it contains depth vocabulary.
    When the name is neutral, credits decide. Semester confirms or adjusts by one notch.
    A 5-credit core CS subject is ALWAYS at least Intermediate.

EXAMPLES:
  "Data Structures" (5 cr, Sem 2)        → Intermediate  (neutral name → credits decide)
  "Digital Logic Design" (5 cr, Sem 2)   → Intermediate  (neutral name → credits decide)
  "Introduction to Communication Systems" (5 cr) → Beginner (explicit Intro overrides credits)
  "Design Thinking Workshop" (2 cr)      → Beginner      (workshop + low credits)
  "Rural Internship" (2.5 cr)            → Beginner      (internship = exposure only)
  "Machine Learning" (4 cr, Sem 6)       → Intermediate  (neutral name → 4 cr + sem 6)
  "Advanced Algorithms" (3 cr, Sem 7)    → Advanced      (name says Advanced)

━━━ SCORE IS FOR PERFORMANCE, NOT PROFICIENCY ━━━
Do NOT use score_percentage to determine inferred_proficiency.
Score data is handled separately by the system.
Tag every entry proficiency_rationale with "No curriculum DB — " so reviewers
know this is an estimate, e.g. "No curriculum DB — subject name neutral, 4 credits → Intermediate".

Subjects to process:
{subjects_json}

Return ONLY a valid JSON array:
[
  {{
    "skill": "Data Structure Implementation",
    "subject": "Data Structures",
    "sub_code": "CSE1101",
    "inferred_proficiency": "Intermediate",
    "proficiency_rationale": "Core CS subject, 5 credits → Intermediate"
  }}
]"""


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

_CODE_PATTERN = re.compile(r'\b[A-Z]{2,5}\d{3,}[A-Z0-9]*\b')
_PUNCT = re.compile(r"[''\"()\[\]{}.,;:\-_/\\]")
_SPACES = re.compile(r'\s+')

def _normalise_subject(s: str) -> str:
    """Strip subject codes, punctuation, lowercase, compress spaces."""
    s = _CODE_PATTERN.sub(' ', s)
    s = _PUNCT.sub(' ', s)
    return _SPACES.sub(' ', s).lower().strip()


def _match_subject(curriculum_subj: str, marksheet_subjects: list[dict]) -> tuple:
    """Return (best_match_dict, ratio) or (None, 0.0)."""
    from difflib import SequenceMatcher

    code_match = _CODE_PATTERN.search(curriculum_subj)
    curr_code = code_match.group(0) if code_match else None
    norm_curr = _normalise_subject(curriculum_subj)
    best, best_ratio = None, 0.0

    for ms in marksheet_subjects:
        ms_code = str(ms.get("sub_code") or "").strip()
        if curr_code and ms_code and curr_code.upper() == ms_code.upper():
            return ms, 1.0

        norm_ms = _normalise_subject(ms.get("subject_name", ""))
        ratio = SequenceMatcher(None, norm_curr, norm_ms).ratio()
        if ratio > best_ratio:
            best_ratio, best = ratio, ms

    return (best, best_ratio) if best_ratio >= 0.72 else (None, 0.0)


_GRADE_MAP_FALLBACK = {"O": 10, "A+": 9, "A": 8, "B+": 7, "B": 6, "C+": 5.5,
                       "C": 5, "D": 4, "P": 4, "F": 0}

def _score_pct(subj: dict, max_gp: float = 10.0) -> float | None:
    """Return score percentage."""
    sp = subj.get("score_percentage")
    if sp is not None:
        try:
            return round(float(sp), 1)
        except (ValueError, TypeError):
            pass

    if subj.get("marks_obtained") is not None and subj.get("marks_total"):
        try:
            return round(float(subj["marks_obtained"]) / float(subj["marks_total"]) * 100, 1)
        except (ValueError, TypeError, ZeroDivisionError):
            pass

    gp = subj.get("gp")
    if gp is not None and max_gp > 0:
        try:
            return round(float(gp) / max_gp * 100, 1)
        except (ValueError, TypeError):
            pass

    grade = str(subj.get("grade") or "").strip().upper()
    if grade in ("N/A", "NA", "I", "R", ""):
        return None
    pts = _GRADE_MAP_FALLBACK.get(grade)
    if pts is not None:
        return round(pts / max_gp * 100, 1)

    return None


def _extract_max_gp(marksheet_data: dict) -> float:
    """Parse the grade_scheme string to find the highest GP value."""
    scheme = str(marksheet_data.get("grade_scheme", ""))
    nums = re.findall(r"=(\d+(?:\.\d+)?)", scheme)
    if nums:
        try:
            return max(float(n) for n in nums)
        except ValueError:
            pass
    return 10.0


# ══════════════════════════════════════════════════════════════════════════════
# GEMINI CALLS
# ══════════════════════════════════════════════════════════════════════════════

def extract_marksheet(text: str) -> dict:
    """Gemini: parse marksheet → student_info + subjects_extracted."""
    return _call_gemini_json(
        prompt=f"{MARKSHEET_PARSE_PROMPT}\n\nMARKSHEET TEXT:\n{text}",
        max_tokens=4096,
    )


def infer_skills_no_curriculum(subjects: list[dict]) -> list[dict]:
    """Gemini: infer skills from subject names when no curriculum DB available."""
    subjects_json = json.dumps([{
        "subject_name": s.get("subject_name", ""),
        "sub_code": s.get("sub_code"),
        "score_percentage": s.get("score_percentage"),
        "credits": s.get("credits"),
    } for s in subjects], indent=2)

    raw = _call_gemini_json(
        prompt=SKILLS_INFER_PROMPT.format(subjects_json=subjects_json),
        max_tokens=4096,
    )
    return raw if isinstance(raw, list) else []


# ══════════════════════════════════════════════════════════════════════════════
# PASSPORT ASSEMBLY
# ══════════════════════════════════════════════════════════════════════════════

def _lightcast_entry_from_row(row: dict) -> dict | None:
    """Build a lightcast_skills list entry from a pre-mapped curriculum DB row."""
    if row.get("lightcast_source") != "verified":
        return None
    lc_skill = row.get("lightcast_skill", "").strip()
    if not lc_skill:
        return None
    return {
        "talent_skill": lc_skill,
        "talent_category": row.get("lightcast_talent_category", ""),
        "talent_subcategory": row.get("lightcast_talent_subcategory", ""),
        "category": row.get("lightcast_category", ""),
        "subcategory": row.get("lightcast_subcategory", ""),
        "similarity": row.get("lightcast_match_confidence", 0.0),
    }


def build_passport_with_curriculum(marksheet_data: dict, curriculum: list[dict]) -> dict:
    """Hybrid matching logic for passport generation."""
    from app.core.database import get_col
    from difflib import SequenceMatcher

    subjects = marksheet_data.get("subjects_extracted", [])
    max_gp = _extract_max_gp(marksheet_data)
    skill_passport = []
    matched_ms_indices: set[int] = set()

    curriculum_by_subj: dict[str, list[dict]] = defaultdict(list)
    for row in curriculum:
        curriculum_by_subj[row.get("subject", "")].append(row)

    # Pass 1: curriculum subjects → match against marksheet
    for curr_subj, rows in curriculum_by_subj.items():
        ms_match, ratio = _match_subject(curr_subj, subjects)

        if ms_match:
            for idx, s in enumerate(subjects):
                if s.get("subject_name") == ms_match.get("subject_name"):
                    matched_ms_indices.add(idx)
                    break

            pct = _score_pct(ms_match, max_gp)
            for row in rows:
                curr_prof = row.get("proficiency", "")
                rationale = row.get("proficiency_rationale", "")

                lc_entry = _lightcast_entry_from_row(row)
                lightcast_skills = [lc_entry] if lc_entry else []

                skill_passport.append({
                    "skill": row["skill"],
                    "original_skill": row.get("original_skill", row["skill"]),
                    "subject": curr_subj,
                    "subject_matched": ms_match.get("subject_name", ""),
                    "match_ratio": round(ratio, 2),
                    "proficiency_status": "verified",
                    "curriculum_proficiency": curr_prof,
                    "effective_proficiency": curr_prof,
                    "proficiency_rationale": rationale,
                    "bloom_level": row.get("bloom_level", ""),
                    "bloom_explanation": row.get("bloom_explanation", ""),
                    "credits": row.get("credits", ""),
                    "total_hours": row.get("total_hours", ""),
                    "job_tags": row.get("job_tags", ""),
                    "grade": ms_match.get("grade"),
                    "gp": ms_match.get("gp"),
                    "marks_obtained": ms_match.get("marks_obtained"),
                    "marks_total": ms_match.get("marks_total"),
                    "score_percentage": pct,
                    "lightcast_skills": lightcast_skills,
                    "lightcast_source": row.get("lightcast_source", ""),
                    "summary_text": (
                        f"{pct:.1f}% in {curr_prof}-level '{row['skill']}' "
                        f"(matched: {ms_match.get('subject_name', '')})"
                    ) if pct is not None else
                        f"Grade found but score not calculable for '{row['skill']}'",
                })

    # Pass 2: unmatched marksheet subjects
    unmatched = [s for idx, s in enumerate(subjects) if idx not in matched_ms_indices]
    if unmatched:
        log.info("Unmatched subjects → LLM inference: %s",
                 [s.get("subject_name") for s in unmatched])
        inferred = infer_skills_no_curriculum(unmatched)

        um_by_name = {_normalise_subject(s.get("subject_name", "")): s for s in unmatched}

        for item in inferred:
            subj_name = item.get("subject", "")
            ms = um_by_name.get(_normalise_subject(subj_name))
            if ms is None:
                ms, _ = _match_subject(subj_name, unmatched)
            pct = _score_pct(ms, max_gp) if ms else None
            grade = str((ms or {}).get("grade") or "").strip().upper()
            if grade in ("N/A", "NA", "I", "R"):
                pct = None
            inferred_prof = item.get("inferred_proficiency", "Unknown")

            skill_passport.append({
                "skill": item.get("skill", ""),
                "original_skill": item.get("skill", ""),
                "subject": subj_name,
                "subject_matched": ms.get("subject_name") if ms else None,
                "match_ratio": 1.0 if ms else 0.0,
                "proficiency_status": "inferred",
                "curriculum_proficiency": inferred_prof,
                "effective_proficiency": inferred_prof,
                "proficiency_rationale": item.get("proficiency_rationale",
                                                "Not in curriculum DB — inferred from subject name"),
                "bloom_level": "",
                "bloom_explanation": "",
                "credits": str((ms or {}).get("credits") or ""),
                "total_hours": "",
                "job_tags": item.get("job_tags", []),
                "grade": (ms.get("grade") if ms else None),
                "gp": (ms.get("gp") if ms else None),
                "marks_obtained": (ms.get("marks_obtained") if ms else None),
                "marks_total": (ms.get("marks_total") if ms else None),
                "score_percentage": pct,
                "lightcast_skills": [],
                "lightcast_source": "unmatched",
                "summary_text": (
                    f"Not in curriculum DB — skills inferred from subject name. "
                    + (f"Score: {pct:.1f}%" if pct is not None else
                       f"Grade: {grade}" if grade and grade not in ("N/A", "NA", "I", "R")
                       else "No score data")
                ),
            })

    return _assemble_passport(marksheet_data, skill_passport)


def build_passport_no_curriculum(marksheet_data: dict) -> dict:
    """No curriculum DB available. Gemini infers skills from subject names."""
    subjects = marksheet_data.get("subjects_extracted", [])
    max_gp = _extract_max_gp(marksheet_data)
    inferred = infer_skills_no_curriculum(subjects)

    skill_passport = []
    for item in inferred:
        subj_name = item.get("subject", "")
        ms, _ratio = _match_subject(subj_name, subjects)
        if ms is None:
            subj_by_name = {_normalise_subject(s.get("subject_name", "")): s for s in subjects}
            ms = subj_by_name.get(_normalise_subject(subj_name))
        pct = _score_pct(ms, max_gp) if ms else None

        grade = str((ms or {}).get("grade") or "").strip().upper()
        if grade in ("N/A", "NA", "I", "R"):
            pct = None

        inferred_prof = item.get("inferred_proficiency", "Unknown")

        skill_passport.append({
            "skill": item.get("skill", ""),
            "original_skill": item.get("skill", ""),
            "subject": subj_name,
            "subject_matched": ms.get("subject_name") if ms else None,
            "match_ratio": 1.0 if ms else 0.0,
            "proficiency_status": "inferred",
            "curriculum_proficiency": inferred_prof,
            "effective_proficiency": inferred_prof,
            "proficiency_rationale": item.get("proficiency_rationale",
                                            "Inferred from subject name — no curriculum DB"),
            "bloom_level": "",
            "bloom_explanation": "",
            "credits": str((ms or {}).get("credits") or "") if ms else "",
            "total_hours": "",
            "job_tags": item.get("job_tags", []),
            "grade": (ms.get("grade") if ms else None),
            "gp": (ms.get("gp") if ms else None),
            "marks_obtained": (ms.get("marks_obtained") if ms else None),
            "marks_total": (ms.get("marks_total") if ms else None),
            "score_percentage": pct,
            "lightcast_skills": [],
            "lightcast_source": "unmatched",
            "summary_text": (
                f"No curriculum DB — skills inferred from subject name. "
                + (f"Score: {pct:.1f}%" if pct is not None else
                   f"Grade: {grade}" if grade and grade not in ('N/A', 'NA', 'I', 'R') else "No score data")
            ),
        })

    return _assemble_passport(marksheet_data, skill_passport)


def _assemble_passport(marksheet_data: dict, skill_passport: list[dict]) -> dict:
    """Compute summary stats and assemble final passport dict."""
    verified = [s for s in skill_passport if s["proficiency_status"] == "verified"]
    inferred = [s for s in skill_passport if s["proficiency_status"] == "inferred"]

    lc_verified_count = sum(1 for s in skill_passport if s.get("lightcast_source") == "verified")

    scored_pool = verified if verified else [s for s in skill_passport if s.get("score_percentage") is not None]
    scores = [s["score_percentage"] for s in scored_pool if s["score_percentage"] is not None]
    avg = round(sum(scores) / len(scores), 1) if scores else None

    def _weight(s):
        pct = s.get("score_percentage") or 0
        try:
            cr = float(str(s.get("credits") or "0").split(":")[0]) or 1
        except ValueError:
            cr = 1
        return pct * cr

    top = sorted(scored_pool, key=_weight, reverse=True)
    bottom = sorted([s for s in scored_pool if (s.get("score_percentage") or 101) < 75],
                    key=lambda s: s.get("score_percentage") or 999)

    return {
        **marksheet_data,
        "skill_passport": skill_passport,
        "summary": {
            "total_skills": len(skill_passport),
            "skills_verified": len(verified),
            "skills_inferred": len(inferred),
            "lightcast_verified_skills": lc_verified_count,
            "average_score_percentage": avg,
            "top_skills": [s["skill"] for s in top[:3]],
            "improvement_areas": [s["skill"] for s in bottom[:3]],
        },
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ══════════════════════════════════════════════════════════════════════════════

PERF_STYLE = {
    "Master": {"bg": "D1FAE5", "fg": "065F46"},
    "Proficient": {"bg": "DBEAFE", "fg": "1E40AF"},
    "Developing": {"bg": "FEF9C3", "fg": "713F12"},
    "Needs Work": {"bg": "FEE2E2", "fg": "7F1D1D"},
    "Unverified": {"bg": "F1F5F9", "fg": "64748B"},
}

def create_passport_excel(passport: dict) -> io.BytesIO:
    wb = openpyxl.Workbook()

    # Sheet 1: Student Info
    ws_info = wb.active
    ws_info.title = "Student Info"
    ws_info.column_dimensions["A"].width = 26
    ws_info.column_dimensions["B"].width = 40

    si = passport.get("student_info", {})
    ws_info["A1"] = "Skill Passport Report"
    ws_info["A1"].font = Font(bold=True, size=16, name="Calibri", color="0F172A")
    ws_info["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws_info["A2"].font = Font(italic=True, size=10, name="Calibri", color="64748B")

    info_rows = [
        ("Name", si.get("name", "")),
        ("Roll / PRN", si.get("roll_no", "")),
        ("Department", si.get("department", "")),
        ("Semester", si.get("semester", "")),
        ("Academic Year", si.get("year", "")),
        ("College", si.get("college", "")),
        ("SGPA", si.get("sgpa", "")),
        ("Result Format", passport.get("result_format", "")),
        ("Overall %", passport.get("overall_percentage", "")),
    ]
    for r, (k, v) in enumerate(info_rows, 4):
        c_k = ws_info.cell(row=r, column=1, value=k)
        c_k.font = Font(bold=True, name="Calibri", size=11)
        c_k.fill = PatternFill("solid", fgColor="F1F5F9")
        c_k.border = thin_border()
        c_v = ws_info.cell(row=r, column=2, value=v)
        c_v.font = Font(name="Calibri", size=11)
        c_v.border = thin_border()

    # Sheet 2: Subject Results
    ws_sub = wb.create_sheet("Subject Results")
    sub_headers = ["Subject", "Code", "Grade", "GP", "Credits", "Marks Obtained", "Marks Total", "Score %"]
    sub_widths = [40, 14, 8, 6, 8, 15, 12, 10]

    for ci, (hdr, w) in enumerate(zip(sub_headers, sub_widths), 1):
        c = ws_sub.cell(row=1, column=ci, value=hdr)
        c.fill = PatternFill("solid", fgColor=HEADER_BG)
        c.font = Font(bold=True, color=HEADER_FG, size=11, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border("1E293B")
        ws_sub.column_dimensions[get_column_letter(ci)].width = w
    ws_sub.row_dimensions[1].height = 28

    for ri, sub in enumerate(passport.get("subjects_extracted", []), 2):
        row_bg = ALT_BG if ri % 2 == 0 else NORMAL_BG
        values = [
            sub.get("subject_name", ""),
            sub.get("sub_code", ""),
            sub.get("grade", ""),
            sub.get("gp", ""),
            sub.get("credits", ""),
            sub.get("marks_obtained", ""),
            sub.get("marks_total", ""),
            sub.get("score_percentage", ""),
        ]
        for ci, val in enumerate(values, 1):
            c = ws_sub.cell(row=ri, column=ci, value=val)
            c.fill = PatternFill("solid", fgColor=row_bg)
            c.font = Font(name="Calibri", size=10)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center", horizontal="center" if ci > 1 else "left")
    ws_sub.freeze_panes = "A2"

    # Additional sheets can be added here (Skill Passport, Summary, etc.)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
